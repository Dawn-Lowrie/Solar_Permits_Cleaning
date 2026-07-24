import openpyxl as xl       #   library openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows      #   imports the dataframe_to_rows function
import pandas as pd         #   library pandas
from datetime import datetime       #   import the datetime function


def extract_contractor_info(data: pd.DataFrame, name: str, trade: str, address1: str, address2: str, phone: str) -> pd.DataFrame:
    '''
    creates a dataframe with all unique contractor information that will be used in Excel for cleaning contractor name data
    '''
    contractor_temp_df = pd.DataFrame()     #   creates a temporary blank dataframe for sorting contractor data to return only unique values
    contractor_info = pd.DataFrame()        #   creates a temporary blank dataframe for the unique vlaues

    contractor_temp_df['contractor_tuple'] = [(x, y, z, w, v) for x, y, z, w, v in zip(data[name], data[trade].fillna(''), data[address1], data[address2], data[phone].fillna(''))]       #   puts the contractor data into a tuple. .fillna('') is used for later Excel implementation
    contractor_info['contractor_tuple'] = contractor_temp_df['contractor_tuple'].unique()       #   contractor_info dataframe is filled with only unique contractor data
    del contractor_temp_df      #   deletes the temporary contractor_temp_df from memory

    contractor_info[name] = contractor_info['contractor_tuple'].apply(lambda x: x[0])       #   unpacks the tuple into their respective columns
    contractor_info[name] = contractor_info['contractor_tuple'].apply(lambda x: x[1])
    contractor_info[name] = contractor_info['contractor_tuple'].apply(lambda x: x[2])
    contractor_info[name] = contractor_info['contractor_tuple'].apply(lambda x: x[3])
    contractor_info[name] = contractor_info['contractor_tuple'].apply(lambda x: x[4])

    contractor_info = contractor_info.drop('contractor_tuple', axis = 1)        #   drops the 'contractor_tuple' column after all information has been separated into their respective columns
    contractor_info = contractor_info.drop(subset = (name))      #   removes any rows where the name column is NA
    contractor_info['contractorname_clean'] = ''        #   adds a blank column 'contractorname_clean' for later use in the Excel file for manually cleaned data

    return contractor_info      #   returns the contractor data


#   took inspiration from code found here: https://realpython.com/openpyxl-excel-spreadsheets-python/#bonus-working-with-pandas

def excel_append_contractors(data: pd.DataFrame, name: str, trade: str, address1: str, address2: str, phone: str, workbook_name: str, worksheet_name: str):
    '''
    a function that adds all or new contractor data to an excel sheet for cleaning.
    as the dataset is updated, new contractors will be added to the excel sheet.
    '''
    contractor_data = extract_contractor_info(data, name, trade, address1, address2, phone)     #   uses the extract_contractor_info function to create a table of all unique contractor information from the dataset
    workbook = xl.load_workbook(filename = workbook_name)       #   loads the excel workbook
    worksheet = workbook[worksheet_name]        #   loads the excel worksheet from the workbook
    existing_values = worksheet.values        #   takes in the existing values from the worksheet

    existing_data = pd.DataFrame(existing_values)        #   converts those values to a pandas dataframe
    existing_data = existing_data.fillna('')      #   fills any NA data to '' to match the format of the NA data in the contractor dataframe. '' was necessary for the is_address1 function, so this method keeps duplicates from slipping through

    if len(existing_data) == 0:        #   if the contractor_info_xl dataframe is empty (no data has been added yet)
        for row in dataframe_to_rows(contractor_data, index = False, header = True):        #   uses the dataframe_to_rows function to iterate across the rows in the contractor data (including the header from the dataframe)
            worksheet.append(row)       #   adds the rows to the worksheet

    else:       #   if there is data in the worksheet
        existing_data.columns = existing_data.iloc[0]     #   promotes the first row to the column names
        existing_data['contractorname_clean'] = ''     #   nulls out the contractorname_clean column so the merge step is successful in removing existing contractor information
        merged_data = pd.merge(contractor_data, existing_data, how = 'left_anti')        #   merges the contractor data with the information from the worksheet where only the information unique to the contractor data is added (removes entries that already exist in the worksheet)
        for row in dataframe_to_rows(merged_data, index = False, header = False):     #   uses the dataframe_to_rows function to iterate across the rows in the contractor data (excluding the header from the dataframe)
            worksheet.append(row)       #   adds the rows to the worksheet

    workbook.save(workbook_name)      #   saves the workbook


def create_protection_sheet(workbook_name: str, worksheet_name: str):
    '''
    a function that takes the name of an excel workbook and worksheet, copies the worksheet data, changes the copied worksheets title to the current time (YYYY-MM-DD, HHMM), and turns on sheet protection so the data cannot be changed
    this function is used after cleaning the data in excel
    '''
    now = datetime.now()        #   saves the datetime when the function is run
    current_date = now.strftime('%m-%d')       #   creates a string using the datetime data (formatted as: 'MM-DD')
    protection_sheet_name = current_date + ' Protection Sheet'      #   worksheet title that includes the current date and time (easier to tell which is the latest version) and that it's a protection sheet

    workbook = xl.load_workbook(filename = workbook_name)       #   loads the excel workbook
    worksheet = workbook[worksheet_name]        #   loads the worksheet in the excel workbook
    worksheet_copy = workbook.copy_worksheet(worksheet)     #   creates a copy of the worksheet data
    worksheet_copy.title = protection_sheet_name     #   changes the copied data's worksheet title to protection_sheet_name
    protection_sheet_actual_name = worksheet_copy.title     #   edge case if function is ran multiple times in one minute, saves the actual title of the worksheet ('YYYY-MM-DD Protection Sheet1 is saved instead of a different file)
    worksheet_copy.protection.sheet = True      #   turns on worksheet protection, so data cannot be altered
    set_latest_protection_sheet(workbook, protection_sheet_actual_name)        #   adds the data to a hidden sheet called 'META DATA' for future use
    workbook.save(workbook_name)        #   saves the changes to the workbook


def set_latest_protection_sheet(workbook_name: str, worksheet_name: str):
    '''
    a function that saves a worksheet name to a hidden worksheet called 'META DATA' in cell B1
    '''
    if 'META DATA' not in workbook_name.sheetnames:     #   if the workbook doesn't have a sheet called 'META DATA'
        meta_sheet = workbook_name.create_sheet('META DATA')        #   creates a sheet called 'META DATA' and sets the current sheet to 'META DATA'
        meta_sheet['A1'] = 'Latest Protection Sheet Name'       #   sets cell A1 to 'Latest Protection Sheet Name' as a table header
        meta_sheet['B1'] = 'Date Added'
    else:       #   if the workbook does have a sheet called 'META DATA'
        meta_sheet = workbook_name['META DATA']     #   sets the current sheet to 'META DATA"
    
    new_row = [[worksheet_name,'=NOW()']]       #   saves the worksheet name and the NOW function to get the current date and time
    for row in new_row:     #   for the data in new_row
        meta_sheet.append(row)      #   append the data to the 'META DATA' sheet
    
    meta_sheet['B'][-1].number_format = 'dd/mm/yy HH:MM'        #   changes the last entry in column B to be formatted as 'dd/mm/yy HH:MM' (returns a large number that needs to be changed to date time to make sense)

    meta_sheet.sheet_state = 'hidden'       #   hides the 'META DATA' sheet to alleviate clutter on the excel workbook
    
def import_protection_sheet_data(workbook_name: str) -> pd.DataFrame:
    '''
    a function that loads the protection sheet data in
    this function is used after a protection sheet was created using create_protection_sheet
    '''
    workbook = xl.load_workbook(filename = workbook_name)       #   loads the excel workbook
    meta_data = workbook['META DATA']       #   loads the 'META DATA' worksheet
    protection_sheet_name = meta_data['A'][-1].value        #   gets the latest protection sheet name from meta data
    protection_sheet = workbook[protection_sheet_name]      #   loads the latest protection sheet

    data_values = protection_sheet.values       #   loads the protection sheet data values
    cols = next(data_values)        #   saves the first row of the protection sheet (the column headers) as cols
    data = list(data_values)        #   saves the remaining data as data

    joining_data = pd.DataFrame(data, columns = cols)       #   creates a pandas dataframe with the data and headers from cols
    joining_data = joining_data.drop(['contractortrade','contractoraddress1','contractoraddress2','contractorphone'], axis = 1)     #   drops all unneccesary columns from the data (those columns were only useful for the cleaning process)
    joining_data = joining_data.drop_duplicates()       #   drops all duplicates from the dataframe (instances where contractorcompanyname was the same but had different addresses, trades, or phone numbers are reduced to only one per each name)
    joining_data.loc[joining_data['contractorname_clean'].isna(), 'contractorname_clean'] = joining_data['contractorcompanyname']       #   fills all empty space in the 'contractorname_clean' column with data from 'contractorcompanyname' (companies with names that didn't need to be changed)
    
    return joining_data     #   returns the joining data

def clean_contractor_names(data: pd.DataFrame, workbook_name: str, primary_key: str, cleaned_column_name: str) -> pd.DataFrame:
    joining_data = import_protection_sheet_data(workbook_name)      #   uses the import_protection_sheet_data function to import the protection sheet of cleaned contractor data as a pandas dataframe

    merged_data = pd.merge(data, joining_data, on = primary_key, how = 'outer')     #   merges the data with the joining_data on the primary_key
    merged_data[primary_key] = merged_data[cleaned_column_name]     #   replaces the unclean primary_key data with cleaned data
    merged_data = merged_data.drop(cleaned_column_name, axis = 1)       #   drops the cleaned_column_name column after the cleaned data has been saved to the primary_key column

    return merged_data      #   returns the cleaned data